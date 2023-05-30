# ###################################################
# Data migration
# ---------------------------------------------------
# Data migration from old CORE to new CORE
# ###################################################

# #####################################
# Imports
# #####################################

# System includes
from datetime import datetime
import pandas as pd
import openpyxl
import xlwt
import os
import re

# Cotown includes
from dbclient import DBClient

# Constants
PAST = datetime.strptime('1900-01-01', '%Y-%m-%d')

# #####################################
# Auxiliary functions
# #####################################

def get_date(string):

  try:
    return datetime.strptime(str(string), '%Y-%m-%d')
  except:
    return PAST

def check_dates(row, tag):

  # Dates
  dates = {
    'Date_from': row['Date_from'],
    'Date_to'  : row['Date_to'],
    'Check_in' : row['Check_in'],
    'Check_out': row['Check_out']
  }

  # Validate from-to
  if dates['Date_from'] == PAST or dates['Date_to'] == PAST:
    return PAST 
  if dates['Date_from'] >= dates['Date_to']:
    return PAST

  # Validate checkin
  if dates['Check_in'] != PAST:
    if dates['Check_in'] < dates['Date_from']:
      dates['Check_in'] = dates['Date_from']
    if dates['Check_in'] > dates['Date_to']:
      dates['Check_in'] = PAST

  # Validate checkout
  if dates['Check_out'] != PAST:
    if dates['Check_out'] > dates['Date_to']:
      dates['Check_out'] = dates['Date_to']
    if dates['Check_out'] < dates['Date_from']:
      dates['Check_out'] = PAST

  # Validate checkin-checkout
  if dates['Check_in'] != PAST and dates['Check_out'] != PAST:
    if dates['Check_out'] < dates['Check_in']:
      dates['Check_in'] = PAST
      dates['Check_out'] = PAST

  # Return requested date
  return dates[tag]
    
def set_status(row):

  # Finished
  if row['Status'] == 'finished':
    return 'finalizada'

  # Request
  if row['Resource_id'] is None:
    return 'solicitud'

  # Dates
  date_from = row['Date_from']
  date_to   = row['Date_to']
  check_in  = row['Check_in']
  check_out = row['Check_out']

  # Finished
  if check_out != PAST:
    if (now - check_out).days > 0:
      return 'finalizada'
  else:
    if (now - date_to).days > 0:
      return 'finalizada'

  # Started
  if check_in != PAST:
    if (now - check_in).days > 0:
      return 'inhouse'
  else:
    if (now - date_from).days > 0:
      return 'inhouse'

  # Confirmed
  return 'confirmada'

def lookup_resource(code, index=0):

  # Fix resource names
  code = code.replace('..', '.').replace('SAT', 'SA')

  # Lookup place
  try:
    id = df_res.loc[df_res[1] == code, index].values[0]
    return id
  except:
    pass
	
  # Lookup room
  code = code[:16]
  try:
    id = df_res.loc[df_res[1] == code, index].values[0]
    return id
  except:
    pass

  return -1


def lookup_customer(email):

  email = email.split('@')[0] + '@test.com'
  email = re.sub(r'[^a-zA-Z0-9-_.@]', '', email)

  try:
    id = df_cus.loc[df_cus[1] == email, 0].values[0]
    return id
  except:
    pass
	
  return -1


# #####################################
# Query
# #####################################

'''
SELECT b.state AS Status, rq.request_type_id AS Booking_channel_id, r.email AS Customer_id, 
	   rq.rental_deposit_amount, rq.rental_deposit_amount_contract, rq.hiring_expense_amount, rq.cleaning_service_amount, 
	   b.FROM AS Date_from, b.TO AS Date_to, bi.checkin AS Check_in, bi.checkout AS Check_out, r.school_id AS School_id, 
	   REPLACE(REPLACE(re.KEY,'_P0','.P'),'_','.') AS Resource_id, b.payment_method_id AS Payment_method_id, rq.comments
FROM booking_resource br 
LEFT JOIN bookings b ON b.id = br.booking_id 
LEFT JOIN booking_information bi ON bi.booking_id = b.id
LEFT JOIN booking_resource_resident brr ON brr.booking_resource_id = br.resource_id
LEFT JOIN requests rq ON rq.id = b.request_id 
LEFT JOIN requesters r ON r.id = rq.requester_id 
LEFT JOIN resources re ON re.id = br.resource_id
LEFT JOIN blocks bl ON bl.booking_id = br.booking_id
WHERE bl.id IS NULL 
ORDER BY br.id;
'''

# ###################################################
# Environment variables
# ###################################################

SERVER   = str(os.environ.get('COTOWN_SERVER'))
DATABASE = str(os.environ.get('COTOWN_DATABASE'))
DBUSER   = str(os.environ.get('COTOWN_DBUSER'))
DBPASS   = str(os.environ.get('COTOWN_DBPASS'))
SSHUSER  = str(os.environ.get('COTOWN_SSHUSER'))
SSHPASS  = str(os.environ.get('COTOWN_SSHPASS'))

# #####################################
# Main
# #####################################

# DB API
dbClient = DBClient(SERVER, DATABASE, DBUSER, DBPASS, SSHUSER, SSHPASS)
dbClient.connect()

# Load customers
dbClient.select('''SELECT id, "Email" FROM "Customer"."Customer"''')
df_cus = pd.DataFrame.from_records(dbClient.fetchall())
print('Clientes...................: ', df_cus.shape[0])

# Load resources
dbClient.select('''SELECT id, "Code", "Resource_type", "Building_id", "Flat_type_id", "Place_type_id" FROM "Resource"."Resource"''')
df_res = pd.DataFrame.from_records(dbClient.fetchall())
print('Recursos...................: ', df_res.shape[0])
print()

# Disconnect
dbClient.disconnect()

# Load data, csv in Excel format
df = pd.read_csv('bookings.in.csv', delimiter=';', encoding='utf-8')
print('Filas originales...........: ', df.shape[0])

# 1. Customer
df['Customer_id'] = df['Customer'].apply(lambda x: lookup_customer(x))
df = df.drop(df.loc[df['Customer_id'] == -1].index)
print('Filas con cliente..........: ', df.shape[0])

# 2. Resource
df['Resource_id'] = df['Resource'].apply(lambda x: lookup_resource(x))
df = df.drop(df.loc[df['Resource_id'] == -1].index)
print('Filas con recurso..........: ', df.shape[0])

# 3. Request
df['Building_id']   = df['Resource'].apply(lambda x: lookup_resource(x, 3))
df['Flat_type_id']  = df['Resource'].apply(lambda x: lookup_resource(x, 4))
df['Place_type_id'] = df['Resource'].apply(lambda x: lookup_resource(x, 5))

# 4. Calculated & cleaned up columns
df['Booking_referral_id'] = 1
df['Second_resident'] = False
df['Lock'] = False
df['Resource_type'] = 'habitacion'
df['Payer_id']  = df['Customer_id']

# 5. Convert to dates
df['Date_from'] = df['Date_from'].apply(lambda x: get_date(x))
df['Date_to']   = df['Date_to'].apply(lambda x: get_date(x))
df['Check_in']  = df['Check_in'].apply(lambda x: get_date(x))
df['Check_out'] = df['Check_out'].apply(lambda x: get_date(x))

# 6. Validate and drop invalid dates
now = datetime.today()
df['Date_from'] = df.apply(lambda row: check_dates(row, 'Date_from'), axis=1)
df['Date_to']   = df.apply(lambda row: check_dates(row, 'Date_to'),   axis=1)
df['Check_in']  = df.apply(lambda row: check_dates(row, 'Check_in'),  axis=1)
df['Check_out'] = df.apply(lambda row: check_dates(row, 'Check_out'), axis=1)
df = df.drop(df.loc[df['Date_from'] == PAST].index)
print('Filas con fechas correctas.: ', df.shape[0])

# 7. Calculate status
df['Status'] = df.apply(lambda row: set_status(row), axis=1)
print(df['Status'])

# 8. Drop columns
df.drop('Customer', axis=1, inplace=True)
df.drop('Resource', axis=1, inplace=True)
df.drop('rental_deposit_amount', axis=1, inplace=True)
df.drop('rental_deposit_amount_contract', axis=1, inplace=True)
df.drop('hiring_expense_amount', axis=1, inplace=True)
df.drop('cleaning_service_amount', axis=1, inplace=True)

# 7. Reindex
df.reset_index(drop=True, inplace=True)
df.insert(0, 'id', range(1, 1 + len(df)))

# Save data to XLSX
file = 'bookings.out.xlsx'
df.to_excel(file, index=False, startrow=1)
wb = openpyxl.load_workbook(file)
sheet = wb.active
sheet['A1'] = 'Booking.Booking'
wb.save(file)

# Convert to XLS
xls_workbook = xlwt.Workbook(encoding='utf-8')
xls_sheet = xls_workbook.add_sheet('Sheet1')
for row_num, row in enumerate(sheet.iter_rows()):
    for col_num, cell in enumerate(row):
        xls_sheet.write(row_num, col_num, cell.value)
xls_workbook.save(file[:-1])