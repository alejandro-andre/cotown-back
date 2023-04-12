# ###################################################
# Imports
# ###################################################

# System includes
from openpyxl import load_workbook
import os
import io

# Cotown includes
from library.dbclient import DBClient
from library.apiclient import APIClient
from library.load_prices import load_prices
from library.load_resources import load_resources


# ###################################################
# Loader function
# ###################################################

def main():

    # ###################################################
    # Environment variables
    # ###################################################

    SERVER   = str(os.environ.get('COTOWN_SERVER'))
    DATABASE = str(os.environ.get('COTOWN_DATABASE'))
    DBUSER   = str(os.environ.get('COTOWN_DBUSER'))
    DBPASS   = str(os.environ.get('COTOWN_DBPASS'))
    GQLUSER  = str(os.environ.get('COTOWN_GQLUSER'))
    GQLPASS  = str(os.environ.get('COTOWN_GQLPASS'))
    SSHUSER  = str(os.environ.get('COTOWN_SSHUSER'))
    SSHPASS  = str(os.environ.get('COTOWN_SSHPASS'))


    # ###################################################
    # GraphQL and DB client
    # ###################################################

    # graphQL API
    apiClient = APIClient(SERVER)
    apiClient.auth(user=GQLUSER, password=GQLPASS)

    # DB API
    dbClient = DBClient(SERVER, DATABASE, DBUSER, DBPASS, SSHUSER, SSHPASS)
    dbClient.connect()


    # ###################################################
    # Main
    # ###################################################

    # Get upload requests
    dbClient.select('SELECT id, "File" FROM "Batch"."Upload" WHERE "Result" IS NULL')
    data = dbClient.fetchall()

    # Loop thru files
    for file in data:

        # Result
        ok = False
        log = ''

        # Get request files
        entity = 'Batch/Upload'
        id = str(file['id'])
        data = apiClient.getFile(id, entity, 'File')

        # Load excel book
        file = io.BytesIO(data.content)
        workbook = load_workbook(filename=file, read_only=True, data_only=True)

        # Log
        log = ''

        # Process each sheet
        for sheet in workbook.sheetnames:

            # Processing
            log += sheet + '\n'
            sql = 'UPDATE "Batch"."Upload" SET "Result"=%s, "Log"=%s WHERE id=%s'
            dbClient.execute(sql, ('Procesando...', '', id))
            dbClient.commit()               

            # Resources
            if sheet == 'Recursos':
                ok, l = load_resources(dbClient, workbook[sheet])

            # Prices
            elif sheet == 'Precios':
                ok, l = load_prices(dbClient, workbook[sheet])

            # Other
            else:
                ok, l = False, 'Error: Tipo de carga desconcida.'

            # Append log
            log += l + '\n'

        # Save result
        sql = 'UPDATE "Batch"."Upload" SET "Result"=%s, "Log"=%s WHERE id=%s'
        dbClient.execute(sql, ('Ok' if ok else 'Error', log, id))
        dbClient.commit()               


# #####################################
# Main
# #####################################

if __name__ == '__main__':

    main()