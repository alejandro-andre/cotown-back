# ###################################################
# Imports
# ###################################################

# System includes
from io import BytesIO
import calendar
import openpyxl

# Logging
import logging
logger = logging.getLogger('COTOWN')


def to_int(x):
  try:
    return int(x)
  except:
    return 0
  

def to_float(x):
  try:
    return float(x)
  except:
    return 0.0



def budget(apiClient):

  # Log
  logger.info('Retrieving budgets...')

  # CSV header
  c = 0
  budget_result = '"id","doc_id","doc_type","booking","date","provider","customer","resource","product","amount","rate","price","data_type","stay_length","discount_type"\n' 

  # Get files
  files = apiClient.call('{ data: Admin_FilesList ( where: { Name: { LIKE: "Budget%" } } ) { id File { name } } }')
  for file in files['data']:

    # Retrieve file
    data = apiClient.getFile(file['id'], 'Admin/Files')
    bytes = BytesIO(data.content)

    # Log
    logger.info('Calculating budget from "' + file['File']['name'] + '"...')

    # Open XLSX and get each sheet
    workbook = openpyxl.load_workbook(bytes, data_only=True)
    for name in workbook.sheetnames:
    
      # Process sheet
      sheet = workbook[name]
      for row in sheet.iter_rows(min_row=5):
        c += 1
        month = str(row[0].value)[:10]

        # Data
        budget = to_float(row[ 2].value)
        uw     = to_float(row[ 3].value)
        line = ['BUD' + str(c), '-', '-', '(budget)', month, '', '', row[1].value, 'Monthly rent', budget, budget, None, 'Budget', '', '' ]
        budget_result += ','.join([f'"{e}"' for e in line]) + '\n'
        line = ['BUW' + str(c), '-', '-', '(uw)', month, '', '', row[1].value, 'Monthly rent', uw, uw, None, 'UW', '', '' ]
        budget_result += ','.join([f'"{e}"' for e in line]) + '\n'

    # Close worksheet
    workbook.close()

  # Save all results to CSV
  with open('csv/income_budget.csv', 'w') as f:
    f.write(budget_result)

  # Log
  logger.info('Done')


def forecast(apiClient):

  # Log
  logger.info('Retrieving forecast...')

  # CSV header
  c = 0
  forecast_result = '"id","doc_id","doc_type","booking","date","provider","customer","resource","product","amount","rate","price","data_type","stay_length","discount_type"\n' 
  occupancy_result = '"id","data_type","resource","date","occupied","sold","occupied_t","sold_t","booking","stay_length"\n'
  beds_result = '"id","data_type","resource","date","beds","beds_c","beds_cnv","beds_pot","beds_pre","beds_cap","available","convertible","val_current","val_residential","val_cosharing"\n'

  # Get files
  files = apiClient.call('{ data: Admin_FilesList ( where: { Name: { LIKE: "Forecast%" } } ) { id File { name } } }')
  for file in files['data']:

    # Retrieve file
    data = apiClient.getFile(file['id'], 'Admin/Files')
    bytes = BytesIO(data.content)

    # Log
    logger.info('Calculating forecast from "' + file['File']['name'] + '"...')

    # Open XLSX and get each sheet
    workbook = openpyxl.load_workbook(bytes, data_only=True)
    for name in workbook.sheetnames:
    
      # Process sheet
      sheet = workbook[name]
      for row in sheet.iter_rows(min_row=5):
        c += 1
        month = str(row[0].value)[:10]
        days = calendar.monthrange(int(month[:4]), int(month[5:7]))[1]

        # Data
        mgmt_fee  = to_float(row[ 2].value)
        beds      = to_float(row[ 3].value)
        beds_c    = to_float(row[ 4].value)
        beds_ad   = to_float(row[ 5].value)
        beds_st   = to_float(row[ 6].value)
        beds_pot  = to_float(row[ 7].value)
        occu      = to_float(row[ 8].value)
        occ_l     = to_float(row[15].value)
        occ_m     = to_float(row[16].value)
        occ_s     = to_float(row[17].value)
        occ_g     = to_float(row[18].value)
        rent_l    = to_float(row[23].value)
        rent_m    = to_float(row[24].value)
        rent_s    = to_float(row[25].value)
        rent_g    = to_float(row[26].value)
        rent_tot  = to_float(row[27].value)
        srvs_tot  = to_float(row[28].value)
        srvs_cln  = to_float(row[29].value)
        bfee      = to_float(row[30].value)
        reinv     = to_float(row[31].value)
        occ_stab  = to_float(row[39].value)
        mpr       = to_float(row[40].value)
        mpr_st    = to_float(row[41].value)
        mpr_pot   = to_float(row[42].value)
        rent_l_ad = (rent_l * beds_ad / beds_c) if beds_c else 0
        rent_m_ad = (rent_m * beds_ad / beds_c) if beds_c else 0
        rent_s_ad = (rent_s * beds_ad / beds_c) if beds_c else 0
        rent_g_ad = (rent_g * beds_ad / beds_c) if beds_c else 0
        mfee      = round((rent_tot + (srvs_tot + srvs_cln) / 1.21) * mgmt_fee, 2)

        # Forecast
        line = ['FRL' + str(c), '-', '-', '(forecast)', month, '', '', row[1].value, 'Monthly rent', rent_l, rent_l, None, 'Forecast', 'LONG', '' ]
        forecast_result += ','.join([f'"{e}"' for e in line]) + '\n'
        line = ['FRM' + str(c), '-', '-', '(forecast)', month, '', '', row[1].value, 'Monthly rent', rent_m, rent_m, None, 'Forecast', 'MEDIUM', '' ]
        forecast_result += ','.join([f'"{e}"' for e in line]) + '\n'
        line = ['FRS' + str(c), '-', '-', '(forecast)', month, '', '', row[1].value, 'Monthly rent', rent_s, rent_s, None, 'Forecast', 'SHORT', '' ]
        forecast_result += ','.join([f'"{e}"' for e in line]) + '\n'
        line = ['FRG' + str(c), '-', '-', '(forecast)', month, '', '', row[1].value, 'Monthly rent', rent_g, rent_g, None, 'Forecast', 'GROUP', '' ]
        forecast_result += ','.join([f'"{e}"' for e in line]) + '\n'
        line = ['FST' + str(c), '-', '-', '(forecast)', month, '', '', row[1].value, 'Periodic cleaning service', srvs_tot, srvs_tot, None, 'Forecast', '', '' ]
        forecast_result += ','.join([f'"{e}"' for e in line]) + '\n'
        line = ['FSC' + str(c), '-', '-', '(forecast)', month, '', '', row[1].value, 'Check-out cleaning services', srvs_cln, srvs_cln, None, 'Forecast', '', '' ]
        forecast_result += ','.join([f'"{e}"' for e in line]) + '\n'
        line = ['FRR' + str(c), '-', '-', '(forecast)', month, '', '', row[1].value, 'Reinvoices', reinv, reinv, None, 'Forecast', '', '' ]
        forecast_result += ','.join([f'"{e}"' for e in line]) + '\n'
        line = ['FBF' + str(c), '-', '-', '(forecast)', month, '', '', row[1].value, 'Membership fee', bfee, bfee, None, 'Forecast', '', '' ]
        forecast_result += ','.join([f'"{e}"' for e in line]) + '\n'
        line = ['FMF' + str(c), '-', '-', '(forecast)', month, '', '', row[1].value, 'Management fee', mfee, mfee, None, 'Forecast', '', '' ]
        forecast_result += ','.join([f'"{e}"' for e in line]) + '\n'

        # Forecast ajusted
        line = ['ARL' + str(c), '-', '-', '(forecast adjusted)', month, '', '', row[1].value, 'Monthly rent', rent_l_ad, rent_l_ad, None, 'Forecast adjusted', 'LONG', '' ]
        forecast_result += ','.join([f'"{e}"' for e in line]) + '\n'
        line = ['ARM' + str(c), '-', '-', '(forecast adjusted)', month, '', '', row[1].value, 'Monthly rent', rent_m_ad, rent_m_ad, None, 'Forecast adjusted', 'MEDIUM', '' ]
        forecast_result += ','.join([f'"{e}"' for e in line]) + '\n'
        line = ['ARS' + str(c), '-', '-', '(forecast adjusted)', month, '', '', row[1].value, 'Monthly rent', rent_s_ad, rent_s_ad, None, 'Forecast adjusted', 'SHORT', '' ]
        forecast_result += ','.join([f'"{e}"' for e in line]) + '\n'
        line = ['ARG' + str(c), '-', '-', '(forecast adjusted)', month, '', '', row[1].value, 'Monthly rent', rent_g_ad, rent_g_ad, None, 'Forecast adjusted', 'GROUP', '' ]
        forecast_result += ','.join([f'"{e}"' for e in line]) + '\n'

        # MPR
        line = ['MPA' + str(c), '-', '-', '(MPR available)', month, '', '', row[1].value, 'Monthly rent', mpr, mpr, None, 'MPR Available', '', '' ]
        forecast_result += ','.join([f'"{e}"' for e in line]) + '\n'
        line = ['MPC' + str(c), '-', '-', '(MPR convertible)', month, '', '', row[1].value, 'Monthly rent', mpr_st, mpr_st, None, 'MPR Convertible', '', '' ]
        forecast_result += ','.join([f'"{e}"' for e in line]) + '\n'
        line = ['MPP' + str(c), '-', '-', '(MPR potential)', month, '', '', row[1].value, 'Monthly rent', mpr_pot, mpr_pot, None, 'MPR Potential', '', '' ]
        forecast_result += ','.join([f'"{e}"' for e in line]) + '\n'

        # Stabilised
        line = ['SPA' + str(c), '-', '-', '(Stabilised available)', month, '', '', row[1].value, 'Monthly rent', mpr * occ_stab, mpr * occ_stab, None, 'Stabilised Available', '', '' ]
        forecast_result += ','.join([f'"{e}"' for e in line]) + '\n'
        line = ['SPC' + str(c), '-', '-', '(Stabilised convertible)', month, '', '', row[1].value, 'Monthly rent', mpr_st * occ_stab, mpr_st * occ_stab, None, 'Stabilised Convertible', '', '' ]
        forecast_result += ','.join([f'"{e}"' for e in line]) + '\n'
        line = ['SPP' + str(c), '-', '-', '(Stabilised potential)', month, '', '', row[1].value, 'Monthly rent', mpr_pot * occ_stab, mpr_pot * occ_stab, None, 'Stabilised Potential', '', '' ]
        forecast_result += ','.join([f'"{e}"' for e in line]) + '\n'

        # Occupancy forecast
        if beds > 0:
          line = ['FOL' + str(c), 'Forecast', row[1].value, month, occ_l * days * occu * beds_c, occ_l * days * occu * beds_c, 0, 0, '', 'LONG']
          occupancy_result += ','.join([f'"{e}"' for e in line]) + '\n'
          line = ['FOM' + str(c), 'Forecast', row[1].value, month, occ_m * days * occu * beds_c, occ_m * days * occu * beds_c, 0, 0, '', 'MEDIUM']
          occupancy_result += ','.join([f'"{e}"' for e in line]) + '\n'
          line = ['FOS' + str(c), 'Forecast', row[1].value, month, occ_s * days * occu * beds_c, occ_s * days * occu * beds_c, 0, 0, '', 'SHORT']
          occupancy_result += ','.join([f'"{e}"' for e in line]) + '\n'
          line = ['FOG' + str(c), 'Forecast', row[1].value, month, occ_g * days * occu * beds_c, occ_g * days * occu * beds_c, 0, 0, '', 'GROUP']
          occupancy_result += ','.join([f'"{e}"' for e in line]) + '\n'
          line = ['FOC' + str(c), 'Forecast', row[1].value, month, beds, beds_c, beds_st, beds_pot, 0, 0, days * beds, '', 0, 0, 0]
          beds_result += ','.join([f'"{e}"' for e in line]) + '\n'
        if beds_st > 0:
          line = ['SOC' + str(c), 'Stabilised', row[1].value, month, days * beds_st * occ_stab, days * beds_st * occ_stab, 0, 0, '', '']
          occupancy_result += ','.join([f'"{e}"' for e in line]) + '\n'
          line = ['SOC' + str(c), 'Stabilised', row[1].value, month, beds_st, beds_st, beds_st, beds_pot, 0, 0, days * beds_st, '', 0, 0, 0]
          beds_result += ','.join([f'"{e}"' for e in line]) + '\n'

    # Close worksheet
    workbook.close()

  # Save all results to CSV
  with open('csv/income_forecast.csv', 'w') as f:
    f.write(forecast_result)

  # Save all results to CSV
  with open('csv/occupancy_forecast.csv', 'w') as f:
    f.write(occupancy_result)
  with open('csv/beds_forecast.csv', 'w') as f:
    f.write(beds_result)

  # Log
  logger.info('Done')