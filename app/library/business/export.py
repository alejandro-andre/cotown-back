# ###################################################
# Imports
# ###################################################

# System includes
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formula.translate import Translator
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
from io import BytesIO
import pandas as pd
import warnings
import json
import os

# Cotown
from library.services.utils import super_flatten

# Logging
import logging
logger = logging.getLogger('COTOWN')


# ###################################################
# Fill excel with JSON
# ###################################################

def fill_sheet(df, columns, sheet):

  # Validations
  if sheet.title.lower() == 'rooming':

    #IdType
    dvIdType = DataValidation(type='list', formula1='=Id_type!$B$3:$B$9999', allow_blank=True)
    dvIdType.showInputMessage = True
    dvIdType.showErrorMessage = True
    sheet.add_data_validation(dvIdType)

    # Gender
    dvGender = DataValidation(type='list', formula1='=Gender!$B$3:$B$9999', allow_blank=True)
    dvGender.showInputMessage = True
    dvGender.showErrorMessage = True
    sheet.add_data_validation(dvGender)

    # Country
    dvCountry = DataValidation(type='list', formula1='=Country!$B$3:$B$9999', allow_blank=True)
    dvCountry.showInputMessage = True
    dvCountry.showErrorMessage = True
    sheet.add_data_validation(dvCountry)

    # Language
    dvLanguage = DataValidation(type='list', formula1='=Language!$B$3:$B$9999', allow_blank=True)
    dvLanguage.showInputMessage = True
    dvLanguage.showErrorMessage = True
    sheet.add_data_validation(dvLanguage)

  # Select and sort columns
  df = df.reindex([item.split(':')[0] for item in columns], axis=1)

  # Copy formats & styles from first data row
  styles = []
  format = []
  start = sheet.max_row
  for c in range(0, df.shape[1]):
    cell = sheet.cell(row=start, column=c+1)
    styles.append(cell._style)
    format.append(cell.value)

  # Recorrer todas las celdas en la hoja
  for row in sheet.iter_rows():
    for cell in row:
      if cell.value and isinstance(cell.value, str) and "(now)" in cell.value:
        cell.value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

  # Write data
  for r, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
    for c in range(0, df.shape[1]):

      # Get cell
      cell = sheet.cell(row = r + start - 2, column = c + 1)

      # Test validation
      if sheet.title.lower() == 'rooming':
        if columns[c] == 'Id_type.Name':
          dvIdType.add(cell)
        if columns[c] == 'Gender.Name':
          dvGender.add(cell)
        if columns[c] == 'Country.Name' or columns[c] == 'Nationality.Name' or columns[c] == 'Origin.Name':
          dvCountry.add(cell)
        if columns[c] == 'Language.Name':
          dvLanguage.add(cell)

      # Copy style
      cell._style = styles[c]

      # List of dicts
      if isinstance(row[c], list):
        try:
          values = []
          for item in row[c]:
            for key in columns[c].split(':')[1:]:
              item = item[key]
            values.append(str(item))
          cell.value = ','.join(values)
        except:
          cell.value = '[ERROR]'

      # Simple value
      else:
        # Format
        if format[c] == 'date':
          try:
            cell.value = datetime.strptime(row[c][:10], '%Y-%m-%d')
          except Exception as e:
            cell.value = row[c]
        elif format[c] == 'datetime':
          cell.value = row[c]
          if row[c]:
            try:
              cell.value = datetime.strptime(row[c].split('.')[0], '%Y-%m-%dT%H:%M:%S')
            except Exception as e:
              pass
        elif format[c] is not None:
          t = Translator(format[c], 'A1')
          cell.value = t.translate_formula(row_delta = r - 2)
        else:
          cell.value = row[c]


# ###################################################
# Export graphql to excel
# ###################################################

def do_export_to_excel(apiClient, dbClient, name, variables=None, external_sql=None):

  # Process variables (convert lists to tuple, for SQL WHERE IN)
  for var in variables:
    if isinstance(variables[var], str):
      if ',' in variables[var]:
        variables[var] = tuple(variables[var].split(','))
 
  # Get template
  fi = open('templates/report/' + name + '.xlsx', 'rb')
  template = BytesIO(fi.read())

  # Open template
  warnings.simplefilter(action='ignore', category=UserWarning)
  wb = load_workbook(filename=BytesIO(template.read()))
  for sheet in wb.sheetnames:

    # Title
    title = sheet.lower().replace('_', '')

    # Get graphQL query
    query = None
    file = 'templates/report/' + name + '.' + title + '.graphql'
    if os.path.exists(file):   
      fi = open(file, 'r')
      query = fi.read()
      fi.close()

    # Get SQL query
    sql = None
    if external_sql is None:
      file = 'templates/report/' + name + '.' + title + '.sql'
      if os.path.exists(file):
        fi = open(file, 'r')
        sql = fi.read()
        fi.close()
    else:
      sql = external_sql

    # Get columns
    if external_sql is None:
      file = 'templates/report/' + name + '.' + title + '.json'
    else:
      file = 'templates/report/' + name + '.json'
    if os.path.exists(file):
      fi = open(file, 'r')
      columns = json.load(fi)
      fi.close()

    # Get graphQL data
    if query:
      result = apiClient.call(query, variables)
      data = result[next(iter(result.keys()))]
      df = pd.DataFrame([super_flatten(d) for d in data])
      fill_sheet(df, columns, wb[sheet])

    # Get SQL data
    if sql:
      try:
        con = dbClient.getconn()
        if variables == {}:
          cur = dbClient.execute(con, sql)
        else:
          cur = dbClient.execute(con, sql, variables)
        desc = [desc[0] for desc in cur.description]
        data = cur.fetchall()
        cur.close()
      except Exception as e:
        logger.error(e)
        con.rollback()
        dbClient.putconn(con)
        return
      dbClient.putconn(con)
      df = pd.DataFrame(data, columns=desc)
      fill_sheet(df, columns, wb[sheet])

  # Save
  virtual_workbook = BytesIO()
  wb.save(virtual_workbook)
  virtual_workbook.seek(0)
  return virtual_workbook