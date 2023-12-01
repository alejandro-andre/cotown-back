# ###################################################
# Imports
# ###################################################

# System includes
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
import pandas as pd
import tempfile

# Cotown includes
from library.services.config import settings
from library.services.dbclient import DBClient

# Logging
import logging
logger = logging.getLogger('COTOWN')


# ###################################################
# Main
# ###################################################

def do_occupancy(dbClient, vars):

  # ###################################################
  # Lambda functions
  # ###################################################

  def nights(resource, date):

    # First and last days of the month
    mfrom = date.replace(day=1)
    mto = date + relativedelta(months=1) - relativedelta(days=1)
   
    # All bookings for the resource
    rows = df_books[df_books['Resource'] == resource]

    # Sum booked nights
    n = 0
    for _, row in rows.iterrows():
      dfrom = row['Date_from']
      dto = row['Date_to']
      if dto >= mfrom and dfrom <= mto:
        n = n + 1 + (min(mto, dto) - max(mfrom, dfrom)).days
    return n


  def available(flat, date):

    # All flat non availability rows
    rows = df_avail[df_avail['Resource_id'] == flat]

    # Check if the date is in between any range
    for _, row in rows.iterrows():
      if row['Date_from'] <= date <= row['Date_to']:
        return 0  # Not available
    return 1  # Available
 
 
  # ###################################################
  # 0. Preparation
  # ###################################################

  # Connect
  logger.info('Start occupancy report')
  dbClient.connect()

  # 0.1 Get all months from the selected range
  # http://localhost:5000/api/v1/occupancy?fdesde=2023-01-01&fhasta=2024-12-31
  start_date = '2023-07-01'
  if vars.get('fdesde'):
    start_date = vars.get('fdesde')
  end_date = (datetime.now() + timedelta(days=366)).strftime('%Y-%m-%d')
  if vars.get('fhasta'):
    end_date = vars.get('fhasta')
  dates = [date.date() for date in pd.date_range(start=start_date, end=end_date, freq='MS')]

  # 0.2 Get all resources: individual rooms and places
  dbClient.select('''
    SELECT b."Name" AS "Building", r."Code" AS "Resource", CASE WHEN r."Resource_type" = 'piso' THEN r.id ELSE r."Flat_id" END AS "Flat_id"
    FROM "Resource"."Resource" r
      INNER JOIN "Building"."Building" b ON b.id = r."Building_id"
    WHERE NOT EXISTS (SELECT id FROM "Resource"."Resource" rr WHERE rr."Code" LIKE CONCAT(r."Code", '.%'))
    ORDER BY r."Code" ASC
    ''')
  columns = [desc[0] for desc in dbClient.sel.description]
  df_res = pd.DataFrame.from_records(dbClient.fetchall(), columns=columns)
  logger.info('Resources retrieved')


  # ###################################################
  # 1. Beds
  # ###################################################

  # 1.1 Get all unavailabilities that affect stock
  dbClient.select('''
    SELECT "Resource_id", "Date_from", "Date_to"
    FROM "Resource"."Resource_availability" ra
    INNER JOIN "Resource"."Resource_status" rs ON rs.id = ra."Status_id"
    WHERE NOT rs."Available"
    ORDER BY "Resource_id"
    ''')
  columns = [desc[0] for desc in dbClient.sel.description]
  df_avail = pd.DataFrame.from_records(dbClient.fetchall(), columns=columns)
  logger.info('Unavailabilities retrieved')

  # 1.2. Check available resources (beds) by month
  df_beds = df_res.copy()
  for dt in dates:
    df_beds[dt] = df_beds.apply(lambda row: available(row['Flat_id'], dt), axis=1)
  df_beds.drop('Flat_id', axis=1, inplace=True)
  df_res.drop('Flat_id', axis=1, inplace=True)
  logger.info('Beds calculated')


  # ###################################################
  # 2. Income
  # ###################################################

  # 2.1. Get all rents
  dbClient.select('''
    SELECT "Building", "Resource", "Date", SUM("Rent") AS "Rent", SUM("Services") AS "Services" 
    FROM (
      -- Rentas B2C (y otras) facturadas
      SELECT bu."Name" AS "Building", r."Code" AS "Resource", i."Code", DATE_TRUNC('month', i."Issued_date") AS "Date", 
        CASE WHEN p."Product_type_id" = 3 THEN il."Amount" ELSE 0 END AS "Rent",
        CASE WHEN p."Product_type_id" <> 3 THEN il."Amount" ELSE 0 END AS "Services"
      FROM "Billing"."Invoice_line" il 
        INNER JOIN "Billing"."Product" p on p.id = il."Product_id" 
        INNER JOIN "Billing"."Invoice" i on i.id = il."Invoice_id" 
        LEFT JOIN "Booking"."Booking" b on b.id = i."Booking_id" 
        LEFT JOIN "Resource"."Resource" r on r.id = b."Resource_id" 
        LEFT JOIN "Building"."Building" bu on bu.id = r."Building_id"  
      WHERE i."Issued" AND i."Bill_type" <> 'recibo' AND i."Booking_group_id" IS NULL
        AND p."Product_type_id" > 2
    UNION
      -- Rentas B2B facturadas
      SELECT bu."Name" AS "Building", (bu."Code" || '-' || b.id) AS "Resource", i."Code", DATE_TRUNC('month', i."Issued_date") AS "Date", 
        CASE WHEN p."Product_type_id" = 3 THEN il."Amount" ELSE 0 END AS "Rent",
        CASE WHEN p."Product_type_id" <> 3 THEN il."Amount" ELSE 0 END AS "Services"
      FROM "Billing"."Invoice_line" il 
        INNER JOIN "Billing"."Product" p on p.id = il."Product_id" 
        INNER JOIN "Billing"."Invoice" i on i.id = il."Invoice_id" 
        LEFT JOIN "Booking"."Booking_group" b on b.id = i."Booking_group_id" 
        LEFT JOIN "Building"."Building" bu on bu.id = b."Building_id"  
      WHERE i."Issued" AND i."Bill_type" <> 'recibo' AND i."Booking_group_id" IS NOT NULL
        AND p."Product_type_id" > 2
    UNION
      -- Rentas B2C no facturadas
      SELECT 
        bu."Name" AS "Building", r."Code" AS "Resource", '', DATE_TRUNC('month', bp."Rent_date") AS "Date",
        CASE 
          WHEN r."Owner_id" = r."Service_id" THEN bp."Rent" + COALESCE(bp."Rent_discount", 0) + bp."Services" + COALESCE(bp."Services_discount", 0)
          ELSE bp."Rent" + COALESCE(bp."Rent_discount", 0)
        END AS "Rent",
        CASE 
          WHEN r."Owner_id" = r."Service_id" THEN 0
          ELSE bp."Services" + COALESCE(bp."Services_discount", 0)
        END AS "Services"
      FROM "Booking"."Booking_price" bp
      INNER JOIN "Booking"."Booking" b ON b.id = bp."Booking_id"
      INNER JOIN "Resource"."Resource" r ON r.id = b."Resource_id"
      INNER JOIN "Building"."Building" bu on bu.id = r."Building_id"
      WHERE bp."Invoice_rent_id" IS NULL AND "Rent_date" > %s
        AND b."Status" IN ('firmacontrato', 'checkinconfirmado', 'contrato','checkin', 'inhouse', 'checkout') 
    UNION
      -- Rentas B2B no facturadas
      SELECT DISTINCT ON (bp.id)
        bu."Name" AS "Building", (bu."Code" || '-' || b.id) AS "Resource", '', DATE_TRUNC('month', bp."Rent_date") AS "Date", 
        CASE 
          WHEN r."Owner_id" = r."Service_id" THEN b."Rooms" * (bp."Rent" + bp."Services")
          ELSE b."Rooms" * bp."Rent"
        END AS "Rent",
        CASE 
          WHEN r."Owner_id" = r."Service_id" THEN 0
          ELSE b."Rooms" * bp."Services"
        END AS "Services"
      FROM "Booking"."Booking_group" b
        INNER JOIN "Booking"."Booking_group_price" bp ON bp."Booking_id" = b.id
        INNER JOIN "Booking"."Booking_rooming" br on b.id = br."Booking_id" 
        INNER JOIN "Building"."Building" bu on bu.id = b."Building_id" 
        INNER JOIN "Resource"."Resource" r on r.id = br."Resource_id"  
      WHERE bp."Invoice_rent_id" IS NULL AND "Rent_date" > %s
        AND b."Status" IN ('grupoconfirmado', 'inhouse') 
    ) AS income
    GROUP BY 1, 2, 3
    ORDER BY 1, 2, 3
  ''', (start_date, start_date))
  columns = [desc[0] for desc in dbClient.sel.description]
  df_bills = pd.DataFrame.from_records(dbClient.fetchall(), columns=columns)
  if len(df_bills.index) > 0:
    df_bills['Date'] = df_bills['Date'].dt.date
    df_bills['Rent'] = pd.to_numeric(df_bills['Rent'], errors='coerce')
    df_bills['Services'] = pd.to_numeric(df_bills['Services'], errors='coerce')
  logger.info('Prices retrieved')

  # http://localhost:5000/api/v1/occupancy?fdesde=2023-10-01&fhasta=2024-12-31

  # 2.2. Pivot rent income by month
  df_rent = df_bills.pivot_table(index=['Building', 'Resource'], columns='Date', values='Rent', aggfunc='sum', fill_value=0.0).reset_index()
  for date in dates:
    if date not in df_rent.columns:
      df_rent[date] = 0.0
  df_rent = df_rent[['Building', 'Resource'] + dates]
  logger.info('Rent calculated')

  # 2.3. Pivot services income by month
  df_services = df_bills.pivot_table(index=['Building', 'Resource'], columns='Date', values='Services', aggfunc='sum', fill_value=0.0).reset_index()
  for date in dates:
    if date  not in df_services.columns:
      df_services[date] = 0.0
  df_services = df_services[['Building', 'Resource'] + dates]
  logger.info('Services calculated')


  # ###################################################
  # 3. Room nights
  # ###################################################

  # 3.1. Get all bookings
  dbClient.select('''
    SELECT r."Code" AS "Resource", b."Date_from", b."Date_to"
    FROM "Booking"."Booking_detail" b
    INNER JOIN (
      SELECT r."Code", r.id
      FROM "Resource"."Resource" r
      INNER JOIN "Building"."Building" b ON b.id = r."Building_id"
      WHERE NOT EXISTS (SELECT id FROM "Resource"."Resource" rr WHERE rr."Code" LIKE CONCAT(r."Code", '.%'))
      ) AS r ON r.id = b."Resource_id"
    WHERE b."Availability_id" IS NULL
    ORDER BY 1, 2, 3
    ''')
  columns = [desc[0] for desc in dbClient.sel.description]
  df_books = pd.DataFrame.from_records(dbClient.fetchall(), columns=columns)
  logger.info('Bookings retrieved')

  # 3.2. Pivot booked room nights by month
  df_night = df_res.copy()
  for dt in dates:
    df_night[dt] = df_night.apply(lambda row: nights(row['Resource'], dt), axis=1)
  logger.info('Nights calculated')


  # ###################################################
  # Final. To excel
  # ###################################################

  # Generate new book from template
  book = load_workbook('templates/report/occupancy.xlsx')
  logger.info('Template loaded')

  # Replicate columns
  for sheet_name in book.sheetnames:

    # Only dashboard sheets
    if 'data-' not in sheet_name:
      sheet = book[sheet_name]

      # One column per date
      for i, _ in enumerate(dates):
        if i > 0:

          # Copy column formulas and styles
          for row in range(1, sheet.max_row+1):
            src = sheet.cell(row=row, column=2)
            if src.data_type == 'f':
              t = Translator(src.value, 'A1')
              formula  = t.translate_formula(col_delta=i)
            else:
              formula = src.value
            dst = sheet.cell(row=row, column=2+i, value=formula)
            dst._style = src._style

      # Today's date
      sheet.cell(row=1, column=2, value=datetime.now().date())

  # Save book
  temp = tempfile.NamedTemporaryFile(suffix='.xlsx').name
  book.save(temp)
  logger.info('Template prepared')

  # Write data
  with pd.ExcelWriter(temp, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:

    # Write sheets
    df_beds.to_excel(writer, sheet_name='data-beds', float_format = "%0.3f", index=False)
    df_night.to_excel(writer, sheet_name='data-nights', float_format = "%0.3f", index=False)
    df_rent.to_excel(writer, sheet_name='data-rent', float_format = "%0.3f", index=False)
    df_services.to_excel(writer, sheet_name='data-services', float_format = "%0.3f", index=False)

    # Hide data sheets
    #?writer.sheets['data-beds'].sheet_state = 'hidden'
    #?writer.sheets['data-nights'].sheet_state = 'hidden'
    #?writer.sheets['data-rent'].sheet_state = 'hidden'
    #?writer.sheets['data-services'].sheet_state = 'hidden'

    # Disconnect
    logger.info('Report finished')
    dbClient.disconnect()

    # return
    return temp