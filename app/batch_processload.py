# ###################################################
# Batch process
# ---------------------------------------------------
# Process excel file load requests
# ###################################################

# ###################################################
# Imports
# ###################################################

# System includes
from openpyxl import load_workbook
import io

# Cotown includes
from library.services.config import settings
from library.services.dbclient import DBClient
from library.services.apiclient import APIClient
from library.business.load_prices import load_prices
from library.business.load_resources import load_resources
from library.business.load_rooming import load_rooming

# Logging
import logging
from logging.handlers import RotatingFileHandler
logger = logging.getLogger('COTOWN')


# ###################################################
# Loader function
# ###################################################

def main():

  # ###################################################
  # Logging
  # ###################################################

  logger.setLevel(settings.LOGLEVEL)
  formatter = logging.Formatter('[%(asctime)s] [%(name)s] [%(module)s] [%(funcName)s/%(lineno)d] [%(levelname)s] %(message)s')
  console_handler = logging.StreamHandler()
  console_handler.setLevel(settings.LOGLEVEL)
  console_handler.setFormatter(formatter)
  file_handler = RotatingFileHandler('log/batch_processload.log', maxBytes=1000000, backupCount=5)
  file_handler.setLevel(settings.LOGLEVEL)
  file_handler.setFormatter(formatter)
  logger.addHandler(console_handler)
  logger.addHandler(file_handler)
  logger.info('Started')


  # ###################################################
  # GraphQL and DB client
  # ###################################################

  # graphQL API
  apiClient = APIClient(settings.SERVER)
  apiClient.auth(user=settings.GQLUSER, password=settings.GQLPASS)

  # DB API
  dbClient = DBClient(
    host=settings.SERVER,
    dbname=settings.DATABASE,
    user=settings.DBUSER,
    password=settings.DBPASS,
    sshuser=settings.SSHUSER,
    sshpassword=settings.get('SSHPASS', None),
    sshprivatekey=settings.get('SSHPKEY', None)
  )
  dbClient.connect()
  con = dbClient.getconn()


  # ###################################################
  # Main
  # ###################################################

  # Get upload requests
  cur = dbClient.execute(con, 'SELECT id, "File" FROM "Batch"."Upload" WHERE "Result" IS NULL')
  data = cur.fetchall()
  cur.close()

  # Loop thru files
  num = 0
  for file in data:

    # Result
    ok = False
    log = ''

    # Get request files
    entity = 'Batch/Upload'
    id = str(file['id'])
    data = apiClient.getFile(id, entity, 'File')

    # Load excel book
    file = io.BytesIO(data.content)
    workbook = load_workbook(filename=file, read_only=True, data_only=True)

    # Log
    log = ''

    # Processing
    sql = 'UPDATE "Batch"."Upload" SET "Result"=%s, "Log"=%s WHERE id=%s'
    dbClient.execute(con, sql, ('Procesando...', '', id))
    con.commit()        

    # Process first sheet
    sheet = workbook.sheetnames[0]

    # Resources
    if sheet == 'Recursos':
      log += sheet + '\n'
      ok, l = load_resources(dbClient, con, workbook[sheet])

    # Prices
    elif sheet == 'Precios':
      log += sheet + '\n'
      ok, l = load_prices(dbClient, con, workbook[sheet])

    # Rooming list
    elif sheet == 'Rooming':
      log += sheet + '\n'
      ok, l = load_rooming(dbClient, con, workbook[sheet])

    # Ignore list
    elif sheet in ('Tarifas', 'Id_type', 'Gender', 'Country', 'Language'):
      ok, l = True, ''

    # Other
    else:
      log += sheet + '\n'
      ok, l = False, 'Error: Tipo de carga desconcida.'

    # Append log
    log += l + '\n'

    # Save result
    sql = 'UPDATE "Batch"."Upload" SET "Result"=%s, "Log"=%s WHERE id=%s'
    dbClient.execute(con, sql, ('Ok' if ok else 'Error', log, id))
    con.commit()        
    num += 1

  # Info
  logger.info('{} files processed'.format(num))


# #####################################
# Main
# #####################################

if __name__ == '__main__':

  main()
  logger.info('Finished')
