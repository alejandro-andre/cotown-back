# ###################################################
# Data migration
# ---------------------------------------------------
# Data migration from old CORE to new CORE
# ###################################################

# #####################################
# Imports
# #####################################

# System includes
from datetime import datetime
import pandas as pd
import openpyxl
import xlwt
import re

# Cotown includes
from library.services.config import settings
from library.services.dbclient import DBClient

# Constants
PAST = datetime.strptime('1900-01-01', '%Y-%m-%d')


# #####################################
# Query
# #####################################

'''
Backoffice:
https://www.3kcoliving.es/backoffice/login
dalarcon@cotown.com
l3$MjE1VHQ69

Base de datos:
https://3kcoliving.es/phpmyadmin/
vanguardstudenthousing
5vA*7aJouT8%
myvanguardstudenthousing
Qi80#GW1AA7N
 
-- 7a. Datos de reserva individuales
SELECT br.id, b.state AS Status, rq.request_type_id AS Booking_channel_id, r.email AS Customer, 
	   rq.rental_deposit_amount, rq.rental_deposit_amount_contract, rq.hiring_expense_amount, rq.cleaning_service_amount, 
	   DATE_FORMAT(b.FROM, "%Y-%m-%d") AS Date_from, DATE_FORMAT(b.to, "%Y-%m-%d") AS Date_to, DATE_FORMAT(bi.checkin, "%Y-%m-%d") AS Check_in, DATE_FORMAT(bi.checkout, "%Y-%m-%d") AS Check_out, 
       r.school_id AS School_id, DATE_FORMAT(rq.created_at, "%Y-%m-%d") AS Request_date, DATE_FORMAT(b.created_at, "%Y-%m-%d") AS Confirmation_date, 
	   REPLACE(REPLACE(re.KEY,'_P0','.P'),'_','.') AS Resource, b.payment_method_id AS Payment_method_id, rq.comments AS Comments
FROM booking_resource br 
LEFT JOIN bookings b ON b.id = br.booking_id 
LEFT JOIN booking_information bi ON bi.booking_id = b.id
LEFT JOIN booking_resource_resident brr ON brr.booking_resource_id = br.resource_id
LEFT JOIN requests rq ON rq.id = b.request_id 
LEFT JOIN requesters r ON r.id = rq.requester_id 
LEFT JOIN resources re ON re.id = br.resource_id
LEFT JOIN blocks bl ON bl.booking_id = br.booking_id
WHERE bl.id IS NULL
AND b.to > '2023-10-01'
AND re.id IS NOT NULL

UNION

SELECT rq.id, rq.state AS Status, rq.request_type_id AS Booking_channel_id, r.email AS Customer,
	   rq.rental_deposit_amount, rq.rental_deposit_amount_contract, rq.hiring_expense_amount, rq.cleaning_service_amount, 
	   DATE_FORMAT(rq.from, "%Y-%m-%d") AS Date_from, DATE_FORMAT(rq.to, "%Y-%m-%d") AS Date_to, NULL AS Check_in, NULL AS Check_out, 
       r.school_id AS School_id, DATE_FORMAT(rq.created_at, "%Y-%m-%d") AS Request_date, NULL AS Confirmation_date, 
	   REPLACE(REPLACE(re.KEY,'_P0','.P'),'_','.') AS Resource, 1 AS Payment_method_id, rq.comments AS Comments
FROM requests rq
LEFT JOIN requesters r ON r.id = rq.requester_id 
LEFT JOIN resources re ON re.id = rq.resource_id
WHERE rq.state IN ('pending', 'accepted')
AND re.id IS NOT NULL
AND rq.to > '2023-10-01';

-- 4. Precios
SELECT br.id AS "Booking_id", CONCAT (y.number, '-', LPAD(m.number, 2, '0'), '-01') AS "Rent_date", cp.amount AS "Rent", cp.amount_cleaning_service AS "Services"
FROM booking_resource br
LEFT JOIN bookings b ON b.id = br.booking_id
LEFT JOIN requests r ON r.id = b.request_id
LEFT JOIN confirmed_prices cp ON br.id = cp.booking_resource_id
INNER JOIN months m ON m.id = cp.month_id 
INNER JOIN years y ON y.id = m.year_id 
WHERE b.to > '2023-10-01'

UNION 

SELECT r.id AS "Booking_id", CONCAT (y.number, '-', LPAD(m.number, 2, '0'), '-01') AS "Rent_date", cp.amount AS "Rent", cp.amount_cleaning_service AS "Services"
FROM requests r
LEFT JOIN confirmed_prices cp ON r.id = cp.request_id
INNER JOIN months m ON m.id = cp.month_id 
INNER JOIN years y ON y.id = m.year_id 
WHERE r.to > '2023-10-01'

ORDER BY 1, 2, 3;
'''


# #####################################
# Auxiliary functions
# #####################################

def get_date(string):

  try:
    return datetime.strptime(str(string), '%Y-%m-%d')
  except:
    pass
  
  try:
    return datetime.strptime(str(string), '%Y-%m-%d %H:%M:%S')
  except:
    pass

  return PAST

def check_dates(row, tag):

  # Dates
  dates = {
    'Date_from': row['Date_from'],
    'Date_to'  : row['Date_to'],
    'Check_in' : row['Check_in'],
    'Check_out': row['Check_out']
  }

  # Validate from-to
  if dates['Date_from'] == PAST or dates['Date_to'] == PAST:
    return PAST 
  if dates['Date_from'] >= dates['Date_to']:
    return PAST

  # Validate checkin
  if dates['Check_in'] != PAST:
    if dates['Check_in'] < dates['Date_from']:
      dates['Check_in'] = dates['Date_from']
    if dates['Check_in'] > dates['Date_to']:
      dates['Check_in'] = PAST

  # Validate checkout
  if dates['Check_out'] != PAST:
    if dates['Check_out'] > dates['Date_to']:
      dates['Check_out'] = dates['Date_to']
    if dates['Check_out'] < dates['Date_from']:
      dates['Check_out'] = PAST

  # Validate checkin-checkout
  if dates['Check_in'] != PAST and dates['Check_out'] != PAST:
    if dates['Check_out'] < dates['Check_in']:
      dates['Check_in'] = PAST
      dates['Check_out'] = PAST

  # Return requested date
  return dates[tag]

def set_status(row):

  # Finished
  if row['Status'] == 'finished':
    return 'finalizada'

  # Pending
  if row['Status'] == 'pending':
    return 'solicitud'

  # Dates
  date_from = row['Date_from']
  date_to   = row['Date_to']
  check_in  = row['Check_in']
  check_out = row['Check_out']

  # Finished
  if check_out != PAST:
    if (now - check_out).days > 0:
      return 'finalizada'
  else:
    if (now - date_to).days > 0:
      return 'finalizada'

  # Started
  if check_in != PAST:
    if (now - check_in).days > 0:
      return 'inhouse'
  else:
    if (now - date_from).days > 0:
      return 'inhouse'

  # Confirmed
  return 'pendientepago'

def lookup_resource(code, index=0):

  # Fix resource names
  code = code.replace('..', '.').replace('SAT', 'SA')

  # Lookup place
  try:
    id = df_res.loc[df_res[1] == code, index].values[0]
    return id
  except:
    pass
	
  # Lookup room
  code = code[:16]
  try:
    id = df_res.loc[df_res[1] == code, index].values[0]
    return id
  except:
    pass

  return -1

def lookup_booking(id):

  # Lookup booking
  try:
    result = df_bookings.loc[df_bookings['id'] == id, 'id'].values[0]
    return result
  except:
    pass
	
  return -1

def lookup_customer(email):

  #email = email.split('@')[0] + '@test.com'
  email = re.sub(r'[^a-zA-Z0-9-_.@]', '', email)

  # Search main email
  try:
    id = df_cus.loc[df_cus[1] == email, 0].values[0]
    return id
  except:
    pass
	
  # Search other emails
  try:
    id = df_cus.loc[df_cus[2].str.contains(email, case=False, na=False), 0].values[0]
    return id
  except:
    pass

  print(email)
  return -1


# #####################################
# Main
# #####################################

# DB API
dbClient = DBClient(
  host=settings.SERVER, 
  dbname=settings.DATABASE, 
  user=settings.DBUSER, 
  password=settings.DBPASS,
  sshuser=settings.SSHUSER, 
  sshpassword=settings.get('SSHPASS', None),
  sshprivatekey=settings.get('SSHPKEY', None)
)
dbClient.connect()
print('ACCESO A BD')

# #####################################
# Get DB info
# #####################################

# Load customers
dbClient.select('''SELECT id, "Email", "Comments" FROM "Customer"."Customer"''')
df_cus = pd.DataFrame.from_records(dbClient.fetchall())
print('Clientes...................: ', df_cus.shape[0])

# Load resources
dbClient.select('''SELECT id, "Code", "Resource_type", "Building_id", "Flat_type_id", "Place_type_id" FROM "Resource"."Resource"''')
df_res = pd.DataFrame.from_records(dbClient.fetchall())
print('Recursos...................: ', df_res.shape[0])

# Disconnect
dbClient.disconnect()

# #####################################
# Bookings
# #####################################

# Load data, csv in Excel format
print('\nRESERVAS')
df_bookings = pd.read_csv('../migration/bookings.in.csv', delimiter=';', encoding='utf-8')
print('Filas originales...........: ', df_bookings.shape[0])

# 1. Customer
df_bookings['Customer_id'] = df_bookings['Customer'].apply(lambda x: lookup_customer(x))
df_bookings = df_bookings.drop(df_bookings.loc[df_bookings['Customer_id'] == -1].index)
print('Filas con cliente..........: ', df_bookings.shape[0])

# 2. Resource
df_bookings['Resource_id'] = df_bookings['Resource'].apply(lambda x: lookup_resource(x))
df_bookings = df_bookings.drop(df_bookings.loc[df_bookings['Resource_id'] == -1].index)

# 3. Request
df_bookings['Building_id']   = df_bookings['Resource'].apply(lambda x: lookup_resource(x, 3))
df_bookings['Flat_type_id']  = df_bookings['Resource'].apply(lambda x: lookup_resource(x, 4))
df_bookings['Place_type_id'] = df_bookings['Resource'].apply(lambda x: lookup_resource(x, 5))

# 4. Calculated & cleaned up columns
df_bookings['Booking_referral_id'] = 1
df_bookings['Second_resident'] = False
df_bookings['Lock'] = False
df_bookings['Resource_type'] = 'habitacion'
df_bookings['Payer_id']  = df_bookings['Customer_id']

# 5. Convert to dates
df_bookings['Date_from'] = df_bookings['Date_from'].apply(lambda x: get_date(x))
df_bookings['Date_to']   = df_bookings['Date_to'].apply(lambda x: get_date(x))
df_bookings['Check_in']  = df_bookings['Check_in'].apply(lambda x: get_date(x))
df_bookings['Check_out'] = df_bookings['Check_out'].apply(lambda x: get_date(x))
df_bookings['Request_date'] = df_bookings['Request_date'].apply(lambda x: get_date(x))
df_bookings['Confirmation_date'] = df_bookings['Confirmation_date'].apply(lambda x: get_date(x))

# 6. Validate and drop invalid dates
now = datetime.today()
df_bookings['Date_from'] = df_bookings.apply(lambda row: check_dates(row, 'Date_from'), axis=1)
df_bookings['Date_to']   = df_bookings.apply(lambda row: check_dates(row, 'Date_to'),   axis=1)
df_bookings['Check_in']  = df_bookings.apply(lambda row: check_dates(row, 'Check_in'),  axis=1)
df_bookings['Check_out'] = df_bookings.apply(lambda row: check_dates(row, 'Check_out'), axis=1)
df_bookings = df_bookings.drop(df_bookings.loc[df_bookings['Date_from'] == PAST].index)
print('Filas con fechas correctas.: ', df_bookings.shape[0])

# 7. Calculate status
df_bookings['Status'] = df_bookings.apply(lambda row: set_status(row), axis=1)

# 8. Drop columns
df_bookings.drop('Customer', axis=1, inplace=True)
df_bookings.drop('Resource', axis=1, inplace=True)
df_bookings.drop('rental_deposit_amount', axis=1, inplace=True)
df_bookings.drop('rental_deposit_amount_contract', axis=1, inplace=True)
df_bookings.drop('hiring_expense_amount', axis=1, inplace=True)
df_bookings.drop('cleaning_service_amount', axis=1, inplace=True)
df_bookings.drop('Payment_method_id', axis=1, inplace=True)

# Save data to XLSX
file = '../migration/bookings.out.xlsx'
df_bookings.to_excel(file, index=False, startrow=1)
wb = openpyxl.load_workbook(file)
sheet = wb.active
sheet['A1'] = 'Booking.Booking'
wb.save(file)

# Convert to XLS
xls_workbook = xlwt.Workbook(encoding='utf-8')
xls_sheet = xls_workbook.add_sheet('Sheet1')
for row_num, row in enumerate(sheet.iter_rows()):
    for col_num, cell in enumerate(row):
        xls_sheet.write(row_num, col_num, cell.value)
xls_workbook.save(file[:-1])

# #####################################
# Prices
# #####################################

# Load data, csv in Excel format
print('\nPRECIOS')
df_prices = pd.read_csv('../migration/prices.in.csv', delimiter=';', encoding='utf-8')
df_prices = df_prices.drop_duplicates(subset=['Booking_id', 'Rent_date'])
print('Filas no duplicadas........: ', df_prices.shape[0])

# 1. Create row
df_prices['Rent_date'] = df_prices['Rent_date'].apply(lambda x: get_date(x))

# 2. Lookup booking
df_prices['Booking_id'] = df_prices['Booking_id'].apply(lambda x: lookup_booking(x))
df_prices = df_prices.drop(df_prices.loc[df_prices['Booking_id'] == -1].index)
print('Filas con reserva..........: ', df_prices.shape[0])

# 3. Round
df_prices['Rent'] = df_prices['Rent'].round(0);
df_prices['Services'] = df_prices['Services'].round(0);

# 4. Reindex
df_prices.insert(0, 'id', range(1, 1 + len(df_prices)))

# Save data to XLSX
file = '../migration/prices.out.xlsx'
df_prices.to_excel(file, index=False, startrow=1)
wb = openpyxl.load_workbook(file)
sheet = wb.active
sheet['A1'] = 'Booking.Booking_price'
wb.save(file)

# Convert to XLS
xls_workbook = xlwt.Workbook(encoding='utf-8')
xls_sheet = xls_workbook.add_sheet('Sheet1')
for row_num, row in enumerate(sheet.iter_rows()):
    for col_num, cell in enumerate(row):
        xls_sheet.write(row_num, col_num, cell.value)
xls_workbook.save(file[:-1])