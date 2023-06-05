# ###################################################
# Data migration
# ---------------------------------------------------
# Data migration from old CORE to new CORE
# ###################################################

# #####################################
# Imports
# #####################################

from datetime import datetime
import pandas as pd
import numpy as np
import openpyxl
import xlwt
import re

# Constants
PAST = datetime.strptime('1900-01-01', '%Y-%m-%d')

# #####################################
# Auxiliary functions
# #####################################

def clean_email(email):

    valid = re.sub(r'[^a-zA-Z0-9-_.@]', '', email)
    if valid != email:
        print(valid, email)
    return valid


def clean_date(date):
    
    if date != '':
        try:
            n = datetime.today()
            d = datetime.strptime(str(date), '%Y-%m-%d')
            if (n.year - d.year > 18) or \
            (n.year - d.year == 18 and n.month > d.month) or \
            (n.year - d.year == 18 and n.month == d.month and n.day > d.day):
                return d
        except Exception as e:
            pass
    return PAST


def nif_nie(dni, type):

    try:
        # Empty id
        if dni is None:
            return None

        # Cntrol chars
        chars = 'TRWAGMYFPDXBNJZSQVHLCKE'

        # To upper case
        dni = dni.upper()

        # Len must be 9
        if len(dni) != 9:
            return type
        
        # Split in number and control digit
        ctl = dni[8]
        dni = dni[:8]

        # NIE?
        ret = 1
        if dni[0] in 'XYZ':
            dni = dni.replace(dni[0], str('XYZ'.find(dni[0])))
            ret = 2

        # Validate control digit
        if len(dni) == len([n for n in dni if n in '1234567890']) and chars[int(dni)%23] == ctl:
            return ret
        
        # Default type
        return type

    except Exception as e:
        print(e)
        return 99

def consolidate(group):

    row = {}

    for column in group.columns:
        values = group[column].values
        non_empty = [value for value in values if pd.notnull(value) and value != '']
        if non_empty:
            row[column] = max(non_empty)
        else:
            row[column] = ''

    return pd.Series(row, index=group.columns)


# #####################################
# Query
# #####################################

'''
SELECT br.id, 
       r.email as "Email", 
       concat(r.name, ' ', r.surname) AS "Name", 
       d.card_type_id + 2 AS "Id_type_id", 
       d.card AS "Document", 
       r.phone as "Phones", 
       d.address as "Address", 
       d.postal_code as "Zip", 
       d.region as "City", 
       d.country_id as "Country_id", 
       r.sex_id as "Gender_id", 
       d.language as "Language_id", 
       r.nationality_id as "Nationality_id", 
       d.birth_date as "Birth_date", 
       r.school_id as "School_id"
FROM requests rq 
LEFT JOIN requesters r ON r.id = rq.requester_id 
LEFT JOIN bookings b ON rq.id = b.request_id 
LEFT JOIN booking_resource br ON b.id = br.booking_id 
LEFT JOIN datasheets d ON d.booking_id = b.id
ORDER BY r.email;
'''

# #####################################
# Main
# #####################################

# Load data, csv in Excel format
print('CLIENTES')
df = pd.read_csv('migration/customers.in.csv', delimiter=';', encoding='utf-8')
print('Filas originales...........: ', df.shape[0])

# 0. Debug. Change email addresses
df['Email'] = df['Email'].str.split('@').str[0] + '@test.com'

# 1. Remove columns
df.drop('id', axis=1, inplace=True)

# 2. Remove rows without mandatory fields
df.dropna(subset=['Name'], inplace=True)
print('Filas con nombre...........: ', df.shape[0])

# 3. Clean fields
df["Email"] = df["Email"].apply(lambda x: clean_email(x))
df["Birth_date"] = df["Birth_date"].apply(lambda x: clean_date(x))

# 4. Consolidate by 'Email'
df = df.groupby(['Email']).apply(consolidate)
print('Filas con diferente email..: ', df.shape[0])

# 5. Drop duplicates in 'Document'
df = df[~(df['Document'].duplicated(keep='last') & (df['Document'] != ''))]
print('Filas sin id duplicado.....: ', df.shape[0])

# 6. Reindex
df.reset_index(drop=True, inplace=True)
df.insert(0, 'id', range(1, 1 + len(df)))

# 7. Calculated & cleaned up columns
df['Name'] = df['Name'].str.title()
df['Type'] = 'persona'
df['User_name'] = 'N' + df['id'].astype(str).str.zfill(6)
df['Comments'] = np.where(df['Language_id'].ne(''), 'Idiomas: ' + df['Language_id'].astype(str), '')
df.drop('Language_id', axis=1, inplace=True)

# 8. NIF/NIE
df['Id_type_id'] = df.apply(lambda row: nif_nie(row['Document'], row['Id_type_id']), axis=1)

# 9. Cut columns length
df['City'] = df['City'].str.slice(0, 50)
df['Zip'] = df['Zip'].str.slice(0, 20)

# Save data to XLSX
file = 'migration/customers.out.xlsx'
df.to_excel(file, index=False, startrow=1)
wb = openpyxl.load_workbook(file)
sheet = wb.active
sheet['A1'] = 'Customer.Customer'
wb.save(file)

# Convert to XLS
xls_workbook = xlwt.Workbook(encoding='utf-8')
xls_sheet = xls_workbook.add_sheet('Sheet1')
for row_num, row in enumerate(sheet.iter_rows()):
    for col_num, cell in enumerate(row):
        xls_sheet.write(row_num, col_num, cell.value)
xls_workbook.save(file[:-1])