# ###################################################
# Imports
# ###################################################

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formula.translate import Translator
from io import BytesIO
import pandas as pd
import json
import os

# Cotown
from library.services.utils import super_flatten

# Logging
import logging
logger = logging.getLogger('COTOWN')


# ###################################################
# Fill excel with JSON
# ###################################################

def fill_sheet(df, columns, sheet):

  # Select and sort columns
  df = df.reindex([item.split(':')[0] for item in columns], axis=1)

  # Copy styles from first data row
  styles = []
  start = sheet.max_row
  for c in range(0, df.shape[1]):
    cell = sheet.cell(row=start, column=c+1)
    styles.append(cell._style)

  # Write data
  for r, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
    for c in range(0, df.shape[1]):

      # Get cell
      cell = sheet.cell(row = r + start - 2, column = c + 1)

      # Copy style
      cell._style = styles[c]

      # Blank column, skip
      if columns[c] == '':
        continue

      # Formula
      elif columns[c][0] == '=':
        t = Translator(columns[c], 'A1')
        cell.value = t.translate_formula(row_delta = r - 2)

      # List of dicts
      elif isinstance(row[c], list):
        try:
          values = []
          for item in row[c]:
            for key in columns[c].split(':')[1:]:
              item = item[key]
            values.append(str(item))
          cell.value = ','.join(values)
        except:
          cell.value = '[ERROR]'

      # Simple value
      else:
        cell.value = row[c]


# ###################################################
# Export graphql to excel
# ###################################################

def query_to_excel(apiClient, dbClient, name, variables=None):

  # Process variables (convert lists to tuple, for SQL WHERE IN)
  for var in variables:
    if isinstance(variables[var], str):
      if ',' in variables[var]:
        variables[var] = tuple(variables[var].split(','))

  # Querys
  query = None
  sql = None

  # Get template
  fi = open('templates/report/' + name + '.xlsx', 'rb')
  template = BytesIO(fi.read())
  
  # Open template
  wb = load_workbook(filename=BytesIO(template.read()))
  for sheet in wb.sheetnames:

    # Get graphQL query
    file = 'templates/report/' + sheet.lower() + '.graphql'
    if os.path.exists(file):    
      fi = open(file, 'r')
      query = fi.read()
      fi.close()

    # Get SQL query
    else:
      file = 'templates/report/' + sheet.lower() + '.sql'
      fi = open(file, 'r')
      sql = fi.read()
      fi.close()

    # Get columns
    fi = open('templates/report/' + sheet.lower() + '.json', 'r')
    columns = json.load(fi)
    fi.close()

    # Get graphQL data
    if query:
      result = apiClient.call(query, variables)
      data = result[next(iter(result.keys()))]
      df = pd.DataFrame([super_flatten(d) for d in data])
      fill_sheet(df, columns, wb[sheet])

    # Get SQL data
    else:
      try:
        dbClient.connect()
        dbClient.select(sql, variables)
        desc = [desc[0] for desc in dbClient.sel.description]
        data = dbClient.fetchall()
      except:
        dbClient.rollback()
        dbClient.disconnect()
        return
      dbClient.disconnect()
      df = pd.DataFrame(data, columns=desc)
      fill_sheet(df, columns, wb[sheet])

  # Save
  virtual_workbook = BytesIO()
  wb.save(virtual_workbook)
  virtual_workbook.seek(0)
  return virtual_workbook