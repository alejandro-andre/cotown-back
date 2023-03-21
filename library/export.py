# ###################################################
# Imports
# ###################################################

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
import pandas as pd


# ###################################################
# Export to excel
# ###################################################

def export_to_excel(apiClient, query, columns, template, variables=None):

    # Call graphQL endpoint
    data = apiClient.call(query, variables)

    # Create dataframe
    df = pd.json_normalize(data[next(iter(data.keys()))])

    print(df)
    print(columns)

    # Select and sort columns
    df = df.reindex(columns, axis=1)

    # Open template
    wb = load_workbook(filename=BytesIO(template.read()))

    # Copy styles from first data row
    styles = []
    start = wb.active.max_row
    for c in range(0, df.shape[1]):
        cell = wb.active.cell(row=start, column=c + 1)
        styles.append(cell._style)

    # Write data
    for r, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for c in range(0, df.shape[1]):
            cell = wb.active.cell(row=r + start - 2, column=c + 1)
            cell.value = row[c]
            cell._style = styles[c]

    # Save
    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)
    return virtual_workbook