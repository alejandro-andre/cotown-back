# ###################################################
# Batch process
# ---------------------------------------------------
# Process excel file load requests
# ###################################################

# ###################################################
# Imports
# ###################################################

# System includes
from openpyxl import load_workbook
import io

# Cotown includes
from library.services.config import settings
from library.services.dbclient import DBClient
from library.services.apiclient import APIClient
from library.business.load_prices import load_prices
from library.business.load_resources import load_resources

# Logging
import logging
from logging.handlers import RotatingFileHandler
logger = logging.getLogger('COTOWN')


# ###################################################
# Loader function
# ###################################################

def main():

  # ###################################################
  # Logging
  # ###################################################

  logger.setLevel(settings.LOGLEVEL)
  formatter = logging.Formatter('[%(asctime)s] [%(name)s] [%(module)s] [%(funcName)s/%(lineno)d] [%(levelname)s] %(message)s')
  console_handler = logging.StreamHandler()
  console_handler.setLevel(settings.LOGLEVEL)
  console_handler.setFormatter(formatter)
  file_handler = RotatingFileHandler('batch_processload.log', maxBytes=1000000, backupCount=5)
  file_handler.setLevel(settings.LOGLEVEL)
  file_handler.setFormatter(formatter)
  logger.addHandler(console_handler)
  logger.addHandler(file_handler)
  logger.info('Started')


  # ###################################################
  # GraphQL and DB client
  # ###################################################

  # graphQL API
  apiClient = APIClient(settings.SERVER)
  apiClient.auth(user=settings.GQLUSER, password=settings.GQLPASS)

  # DB API
  dbClient = DBClient(settings.SERVER, settings.DATABASE, settings.DBUSER, settings.DBPASS, settings.SSHUSER, settings.SSHPASS)
  dbClient.connect()


  # ###################################################
  # Main
  # ###################################################

  # Get upload requests
  dbClient.select('SELECT id, "File" FROM "Batch"."Upload" WHERE "Result" IS NULL')
  data = dbClient.fetchall()

  # Loop thru files
  num = 0
  for file in data:

    # Result
    ok = False
    log = ''

    # Get request files
    entity = 'Batch/Upload'
    id = str(file['id'])
    data = apiClient.getFile(id, entity, 'File')

    # Load excel book
    file = io.BytesIO(data.content)
    workbook = load_workbook(filename=file, read_only=True, data_only=True)

    # Log
    log = ''

    # Process only first sheet
    sheet = workbook.sheetnames[0]

    # Processing
    log += sheet + '\n'
    sql = 'UPDATE "Batch"."Upload" SET "Result"=%s, "Log"=%s WHERE id=%s'
    dbClient.execute(sql, ('Procesando...', '', id))
    dbClient.commit()         

    # Resources
    if sheet == 'Recursos':
      ok, l = load_resources(dbClient, workbook[sheet])

    # Prices
    elif sheet == 'Precios':
      ok, l = load_prices(dbClient, workbook[sheet])

    # Other
    else:
      ok, l = False, 'Error: Tipo de carga desconcida.'

    # Append log
    log += l + '\n'

    # Save result
    sql = 'UPDATE "Batch"."Upload" SET "Result"=%s, "Log"=%s WHERE id=%s'
    dbClient.execute(sql, ('Ok' if ok else 'Error', log, id))
    dbClient.commit()         
    num += 1

  # Info
  logger.info('{} files processed'.format(num))


# #####################################
# Main
# #####################################

if __name__ == '__main__':

  main()
  logger.info('Finished')
